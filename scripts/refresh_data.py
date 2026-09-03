#!/usr/bin/env python3
"""
Pulls the current school year's Incidents workbooks from the district's Google Drive
folder and regenerates data.json for the dashboard.

Requires a Google Cloud service account with read access ("Viewer") to the two
district Drive folders. The service account's JSON key is provided via the
GOOGLE_SERVICE_ACCOUNT_JSON environment variable (set as a GitHub Actions secret).

Set these as repo variables/secrets (see README.md for full setup):
  DRIVE_FOLDER_ID_THIS_YEAR   - folder ID for the current school year (e.g. "26-27")
  DRIVE_FOLDER_ID_LAST_YEAR   - folder ID for last school year (e.g. "25-26"), only
                                 needed the first time -- once baked into data.json,
                                 last year's numbers don't need to be re-pulled every run.
"""
import os
import io
import json
import re
import sys
from datetime import datetime, timezone

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google.oauth2 import service_account
import openpyxl

SCHOOLS = ["McHarg Elementary", "Belle Heth Elementary", "Dalton Intermediate", "Radford High"]
MONTHS = ["August","September","October","November","December","January","February","March","April","May","June"]
CATEGORY_KEYS = ["BAP", "BESO", "BSC", "BSO", "RB"]

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


def get_drive_service():
    creds_json = os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"]
    info = json.loads(creds_json)
    creds = service_account.Credentials.from_service_account_info(info, scopes=SCOPES)
    return build("drive", "v3", credentials=creds)


def list_children(service, folder_id):
    results = []
    page_token = None
    while True:
        resp = service.files().list(
            q=f"'{folder_id}' in parents and trashed = false",
            fields="nextPageToken, files(id, name, mimeType)",
            pageToken=page_token,
        ).execute()
        results.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return results


def download_xlsx(service, file_id):
    request = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf


def parse_incidents_sheet(wb):
    """Parses the 'Incidents' sheet's Total:, category, and School Total Incidents: rows.

    IMPORTANT: labels are NOT always in column A. The "School Total Incidents:" row in
    particular is indented several columns in. So instead of assuming row[0] is the
    label, we scan every cell in the row for a matching label, then read the numeric
    values that follow it in that same row.
    """
    ws = wb["Incidents"] if "Incidents" in wb.sheetnames else wb.worksheets[0]
    district_total = None
    school_totals = None
    categories = {}

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue

        for i, cell in enumerate(row):
            if cell is None:
                continue
            label = str(cell).strip()
            if not label:
                continue

            if label.startswith("Total:"):
                nums = [c for c in row[i + 1:] if isinstance(c, (int, float))]
                if nums:
                    district_total = int(nums[0])
                break

            if "School Total Incidents" in label:
                nums = [int(c) for c in row[i + 1:] if isinstance(c, (int, float))]
                if len(nums) >= 4:
                    school_totals = nums[:4]
                break

            matched_category = False
            for key in CATEGORY_KEYS:
                if label.startswith(key):
                    nums = [c for c in row[i + 1:] if isinstance(c, (int, float))]
                    if nums:
                        categories[key] = int(nums[0])
                    matched_category = True
                    break
            if matched_category:
                break

    return district_total, school_totals, categories


def find_month_workbook(service, month_folder_id, year_label, month_name):
    children = list_children(service, month_folder_id)
    for f in children:
        name = f["name"]
        if "Incident" in name and f["mimeType"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            return f
    return None


def find_month_folders(service, year_folder_id):
    """Maps month name -> folder id, matching folder names like '1. August'."""
    children = list_children(service, year_folder_id)
    mapping = {}
    for f in children:
        if f["mimeType"] != "application/vnd.google-apps.folder":
            continue
        for month in MONTHS:
            if month.lower() in f["name"].lower():
                mapping[month] = f["id"]
    return mapping


def build_year_data(service, year_folder_id, year_label, existing=None):
    """Pulls all available months for a school year folder. `existing` lets us
    keep previously-fetched months (e.g. last year, already complete) without
    re-downloading every run."""
    data = dict(existing or {})
    month_folders = find_month_folders(service, year_folder_id)

    for month, folder_id in month_folders.items():
        wb_file = find_month_workbook(service, folder_id, year_label, month)
        if not wb_file:
            continue
        buf = download_xlsx(service, wb_file["id"])
        wb = openpyxl.load_workbook(buf, data_only=True)
        district_total, school_totals, categories = parse_incidents_sheet(wb)
        if district_total is None or school_totals is None:
            print(f"WARNING: could not parse {year_label} {month}, skipping", file=sys.stderr)
            continue
        data[month] = {
            "district": district_total,
            "schools": school_totals,
            "categories": categories,
            "flagged": data.get(month, {}).get("flagged", False),  # preserve manual flags
        }
    return data


def main():
    service = get_drive_service()

    this_year_folder = os.environ["DRIVE_FOLDER_ID_THIS_YEAR"]
    last_year_folder = os.environ.get("DRIVE_FOLDER_ID_LAST_YEAR")

    # Load existing data.json (if present) so we don't lose manual "flagged" markers
    # or last year's data if DRIVE_FOLDER_ID_LAST_YEAR isn't set on every run.
    existing = {}
    if os.path.exists("data.json"):
        with open("data.json") as f:
            existing = json.load(f).get("years", {})

    years_out = {}
    years_out["26-27"] = build_year_data(service, this_year_folder, "26-27", existing.get("26-27"))
    if last_year_folder:
        years_out["25-26"] = build_year_data(service, last_year_folder, "25-26", existing.get("25-26"))
    else:
        years_out["25-26"] = existing.get("25-26", {})

    output = {
        "schools": SCHOOLS,
        "months": MONTHS,
        "lastUpdated": datetime.now(timezone.utc).isoformat(),
        "years": years_out,
    }

    with open("data.json", "w") as f:
        json.dump(output, f, indent=2)

    print("data.json refreshed successfully.")


if __name__ == "__main__":
    main()
